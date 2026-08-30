#!/usr/bin/env python3
"""Compact workbook inspector and batch writer for the market-report framework.

The report sheet declares roughly 22,000 cells but fills fewer than 300 of them,
so an unguarded dump costs orders of magnitude more than the content is worth.
Both subcommands exist to keep that cost bounded and predictable.

    python wb.py inspect <file.xlsx> [--sheet NAME]... [--max-chars N]
    python wb.py write   <file.xlsx> <cells.json>
    python wb.py selftest
"""

import argparse
import json
import sys
import tempfile
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter


def merged_anchors(ws):
    """Map every cell inside a merged range to that range's top-left anchor."""
    anchors = {}
    for rng in ws.merged_cells.ranges:
        anchor = f"{get_column_letter(rng.min_col)}{rng.min_row}"
        for row in range(rng.min_row, rng.max_row + 1):
            for col in range(rng.min_col, rng.max_col + 1):
                anchors[f"{get_column_letter(col)}{row}"] = anchor
    return anchors


def inspect(path, sheets, max_chars):
    wb = openpyxl.load_workbook(path)
    for ws in wb:
        if sheets and ws.title not in sheets:
            continue
        ranges = sorted(str(r) for r in ws.merged_cells.ranges)
        print(f"# sheet: {ws.title} | dims {ws.dimensions} | merged {len(ranges)}")
        if ranges:
            print("# merged: " + " ".join(ranges))
        count = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                text = str(cell.value)
                if len(text) > max_chars:
                    text = f"{text[:max_chars]}…(+{len(text) - max_chars} chars)"
                print(f"{cell.coordinate}\t" + text.replace("\n", "\\n"))
                count += 1
        print(f"# non-empty: {count}\n")


def write(path, payload):
    """Write {sheet: {cell: value}}. Validates everything before saving anything."""
    wb = openpyxl.load_workbook(path)
    errors = []
    planned = []
    for sheet, cells in payload.items():
        if sheet not in wb.sheetnames:
            errors.append(f"no such sheet: {sheet!r} (have {wb.sheetnames})")
            continue
        ws = wb[sheet]
        anchors = merged_anchors(ws)
        for coord, value in cells.items():
            anchor = anchors.get(coord)
            if anchor and anchor != coord:
                errors.append(
                    f"{sheet}!{coord} is inside a merged range; write {anchor} instead"
                )
                continue
            planned.append((ws, coord, value))

    if errors:
        for err in errors:
            print(f"ERROR: {err}", file=sys.stderr)
        raise SystemExit(1)

    for ws, coord, value in planned:
        ws[coord] = value
    wb.save(path)
    print(f"wrote {len(planned)} cells to {path}")


def selftest():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.merge_cells("B2:C3")
    ws["A1"] = "keep"
    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "t.xlsx"
        wb.save(path)

        write(path, {"S": {"A2": "value", "B2": 42}})
        check = openpyxl.load_workbook(path)["S"]
        assert check["A1"].value == "keep", "existing cell was clobbered"
        assert check["A2"].value == "value"
        assert check["B2"].value == 42, "merged anchor write failed"

        try:
            write(path, {"S": {"C3": "bad"}})
        except SystemExit:
            pass
        else:
            raise AssertionError("writing a non-anchor merged cell must fail")

        try:
            write(path, {"S": {"A5": "ok"}, "Nope": {"A1": "x"}})
        except SystemExit:
            pass
        else:
            raise AssertionError("unknown sheet must fail")
        assert openpyxl.load_workbook(path)["S"]["A5"].value is None, (
            "a rejected batch must write nothing at all"
        )
    print("selftest ok")


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    sub = parser.add_subparsers(dest="cmd", required=True)

    p = sub.add_parser("inspect", help="dump non-empty cells and merged ranges")
    p.add_argument("file")
    p.add_argument("--sheet", action="append", default=[])
    p.add_argument("--max-chars", type=int, default=1000)

    p = sub.add_parser("write", help="write a {sheet: {cell: value}} JSON batch")
    p.add_argument("file")
    p.add_argument("json")

    sub.add_parser("selftest", help="run the built-in checks")

    args = parser.parse_args()
    if args.cmd == "inspect":
        inspect(args.file, set(args.sheet), args.max_chars)
    elif args.cmd == "write":
        write(args.file, json.loads(Path(args.json).read_text(encoding="utf-8")))
    else:
        selftest()


if __name__ == "__main__":
    main()
